# openpyxl 설치필요
import json
import os
import regex as re
from tqdm import tqdm
import time

from langchain.chat_models import ChatOpenAI
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.vectorstores import pinecone
from langchain.embeddings.openai import OpenAIEmbeddings
from langchain_core.documents import Document

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pinecone import Pinecone


with open('apikeys.json', 'r') as f:
    API_KEYS = json.load(f)

# langchain과 pincone을 연동하기위해 환경변수에 API KEY 저장
os.environ['PINECONE_API_KEY'] = API_KEYS['pinecone']['key']
os.environ['PINECONE_API_ENV'] = API_KEYS['pinecone']['env']

PINECONE_INDEX_NAME = 'llm-rag-openai' # 파인콘 인덱스 이름
XLSX_PATH = "C:/Users/rladn/Downloads/case.xlsx" # xlsx파일 경로
CHUNK_SIZE = 1000 # 청크 사이즈
OVERLAP_RATIO = 0.3 # 오버랩 비율
CHUNK_BATCH = 1000 # 최대 청크 배치 (chunks변수 메모리 과다사용 방지)


if __name__ == '__main__':
    # 임베더 (openai ada 002)
    embedder = OpenAIEmbeddings(
        model='text-embedding-ada-002',
        openai_api_key=API_KEYS['open_ai']['key']
    )

    # 파인콘 연결
    pc = Pinecone(
        api_key=os.getenv('PINECONE_API_KEY'),
        environment=os.getenv('PINECONE_API_ENV')
    )
    pc_index = pc.Index(PINECONE_INDEX_NAME)

    # 랭체인과 파인콘 연동
    vectorstore = pinecone.Pinecone(
        index=pc_index,
        embedding=embedder,
        text_key='text'
    )

    # 청크 생성기
    text_splitter=RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE, # 토큰수가 꽤 여유로우므로 size = 1000이어도 좋을것 같다.
        chunk_overlap=int(CHUNK_SIZE * OVERLAP_RATIO),
        length_function=len,
        is_separator_regex=False
    )

    t_start = time.time()
    # openpyxl로 xlsx열고 row별로 작업 (lazy execution)
    workbook = load_workbook(filename=XLSX_PATH)
    worksheet = workbook['Sheet1']
    chunks = []
    chunks_done = 0
    iter = worksheet.iter_rows(min_row=2, max_col=5, values_only=False)
    for row in tqdm(iter, desc='변환 작업중...'):
        # 임베딩 여부 컬럼에 값이 있으면 넘어가기
        # 없으면 True 입력하고 임베딩 진행
        if row[3].value != None:
            continue
        # 사건번호 뽑기 (없으면 null) 
        text = row[1].value
        match = re.search(r"\d+[가-힣]*\s*\d+(\(전합\))?[,\s]", text)
        case_num = text[:match.end()-1] if match else "null"

        # 임베딩할 텍스트 뽑고 선고년도 추출 (없으면 -1)
        text = BeautifulSoup(row[2].value, 'html.parser').get_text(strip=True)
        match = re.search(r'(선고|회시|회신)\D*\d{4}', text)
        year = int(text[match.end()-4:match.end()]) if match else -1 

        # 도큐먼트화, 이후 청크로 분리 (사건번호를 prefix로 추가)
        doc = Document(page_content=text, metadata={'case_num':case_num, 'year':year})
        docs = text_splitter.split_documents([doc])
        for i in range(len(docs)):
            docs[i].page_content = case_num + f' 분할 {i} : ' + docs[i].page_content
        chunks.extend(docs)
        
        # 임베딩 완료표시
        row[3].value = 'T'
        # 청크 갯수가 배치 숫자를 넘기면 파인콘에 upsert
        if len(chunks) > CHUNK_BATCH:
            vectorstore.add_documents(chunks)
            chunks_done += len(chunks)
            chunks = []
    
    # workbook 종료 및 저장
    workbook.save(filename=XLSX_PATH)

    # 남은 chunk 마저 업로드
    if len(chunks) >= 1:
        vectorstore.add_documents(chunks)
        chunks_done += len(chunks)

    t_end = time.time()
    print(f"####작업완료####\n업로드된 chunk : {chunks_done}개\n소요시간 : {t_end-t_start:.2f}초")